# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ------------------------------------------------------------------------------
# Name:         Isogeo to Microsoft Excel 2010
# Purpose:      Get metadatas from an Isogeo share and store it into
#               a Excel worksheet. It's one of the submodules of
#               isogeo2office (https://bitbucket.org/isogeo/isogeo-2-office).
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      30/01/2016
# ------------------------------------------------------------------------------

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from datetime import datetime

# 3rd party library
from dateutil.parser import parse as dtparse
from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties

# ##############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2xlsx(Workbook):
    """
    docstring for Isogeo2xlsx
    """
    li_cols_vector = [
                        "Titre",
                        "Nom",
                        "Résumé",
                        "Emplacement",
                        "Groupe de travail",
                        "Mots-clés",
                        "Thématique(s) INSPIRE",
                        "Conformité INSPIRE",
                        "Contexte de collecte",
                        "Méthode de collecte",
                        "Début de validité",
                        "Fin de validité",
                        "Fréquence de mise à jour",
                        "Commentaire",
                        "Création",
                        "# mises à jour",
                        "Dernière mise à jour",
                        "Publication",
                        "Format (version - encodage)",
                        "SRS (EPSG)",
                        "Emprise",
                        "Géométrie",
                        "Résolution",
                        "Echelle",
                        "# Objets",
                        "# Attributs",
                        "Attributs (A-Z)",
                        "Spécifications",
                        "Cohérence topologique",
                        "Conditions",
                        "Limitations",
                        "# Contacts",
                        "Points de contact",
                        "Autres contacts",
                        "Téléchargeable",
                        "Visualisable",
                        "Autres",
                        "Editer",
                        "Consulter",
                        "MD - ID",
                        "MD - Création",
                        "MD - Modification",
                        "MD - Langue",
                    ]

    li_cols_raster = [
                        "Titre",
                        "Nom",
                        "Résumé",
                        "Emplacement",
                        "Groupe de travail",
                        "Mots-clés",
                        "Thématique(s) INSPIRE",
                        "Conformité INSPIRE",
                        "Contexte de collecte",
                        "Méthode de collecte",
                        "Début de validité",
                        "Fin de validité",
                        "Fréquence de mise à jour",
                        "Commentaire",
                        "Création",
                        "# mises à jour",
                        "Dernière mise à jour",
                        "Publication",
                        "Format (version - encodage)",
                        "SRS (EPSG)",
                        "Emprise",
                        "Résolution",
                        "Echelle",
                        "Attributs (A-Z)",
                        "Spécifications",
                        "Cohérence topologique",
                        "Conditions",
                        "Limitations",
                        "# Contacts",
                        "Points de contact",
                        "Autres contacts",
                        "Téléchargeable",
                        "Visualisable",
                        "Autres",
                        "Editer",
                        "Consulter",
                        "MD - ID",
                        "MD - Création",
                        "MD - Modification",
                        "MD - Langue",
                    ]

    li_cols_service = [
                        "Titre",
                        "Nom",
                        "Résumé",
                        "Emplacement",
                        "Groupe de travail",
                        "Mots-clés",
                        "Conformité INSPIRE",
                        "Création",
                        "# mises à jour",
                        "Dernière mise à jour",
                        "Publication",
                        "Format (version)",
                        "Emprise",
                        "Spécifications",
                        "Conditions",
                        "Limitations",
                        "# Contacts",
                        "Points de contact",
                        "Autres contacts",
                        "Téléchargeable",
                        "Visualisable",
                        "Autres",
                        "Editer",
                        "Consulter",
                        "MD - ID",
                        "MD - Création",
                        "MD - Modification",
                        "MD - Langue",
                        ]

    li_cols_resource = [
                        "Titre",
                        "Nom",
                        "Résumé",
                        "Emplacement",
                        "Groupe de travail",
                        "Mots-clés",
                        "Création",
                        "# mises à jour",
                        "Dernière mise à jour",
                        "Publication",
                        "Format (version)",
                        "Conditions",
                        "Limitations",
                        "# Contacts",
                        "Points de contact",
                        "Autres contacts",
                        "Téléchargeable",
                        "Visualisable",
                        "Autres",
                        "Editer",
                        "Consulter",
                        "MD - ID",
                        "MD - Création",
                        "MD - Modification",
                        "MD - Langue",
                    ]

    def __init__(self):
        """ Isogeo connection parameters

        Keyword arguments:

        """
        super(Isogeo2xlsx, self).__init__()
        # super(Isogeo2xlsx, self).__init__(write_only=True)

        # deleting the default worksheet
        ws = self.active
        self.remove_sheet(ws)

    def md2wb(self, wbsheet, offset, li_mds, li_catalogs):
        """
        parses Isogeo metadatas and write it into the worksheet
        """
        # looping on metadata
        for md in li_mds:
            # incrementing line number
            offset += 1
            # extracting & parsing tags
            tags = md.get("tags")
            li_motscles = []
            li_theminspire = []
            srs = ""
            owner = ""
            inspire_valid = 0
            # looping on tags
            for tag in tags.keys():
                # free keywords
                if tag.startswith('keyword:isogeo'):
                    li_motscles.append(tags.get(tag))
                    continue
                else:
                    pass
                # INSPIRE themes
                if tag.startswith('keyword:inspire-theme'):
                    li_theminspire.append(tags.get(tag))
                    continue
                else:
                    pass
                # workgroup which owns the metadata
                if tag.startswith('owner'):
                    owner = tags.get(tag)
                    continue
                else:
                    pass
                # coordinate system
                if tag.startswith('coordinate-system'):
                    srs = tags.get(tag)
                    continue
                else:
                    pass
                # format pretty print
                if tag.startswith('format:'):
                    format_lbl = tags.get(tag)
                    continue
                else:
                    format_lbl = "NR"
                    pass
                # INSPIRE conformity
                if tag.startswith('conformity:inspire'):
                    inspire_valid = 1
                    continue
                else:
                    pass

            # HISTORY ###########
            if md.get("created"):
                data_created = dtparse(md.get("created")).strftime("%a %d %B %Y")
            else:
                data_created = "NR"
            if md.get("modified"):
                data_updated = dtparse(md.get("modified")).strftime("%a %d %B %Y")
            else:
                data_updated = "NR"
            if md.get("published"):
                data_published = dtparse(md.get("published")).strftime("%a %d %B %Y")
            else:
                data_published = "NR"

            # formatting links to visualize on OpenCatalog and edit on APP
            link_visu = 'HYPERLINK("{0}"; "{1}")'.format(url_OpenCatalog + "/m/" + md.get('_id'),
                                                         "Visualiser")
            link_edit = 'HYPERLINK("{0}"; "{1}")'.format("https://app.isogeo.com/resources/" + md.get('_id'),
                                                         "Editer")
            # format version
            if md.get("formatVersion"):
                format_version = u"{0} ({1} - {2})".format(format_lbl,
                                                           md.get("formatVersion"),
                                                           md.get("encoding"))
            else:
                format_version = format_lbl

            # formatting contact details
            contacts = md.get("contacts")
            if len(contacts):
                contacts_cct = ["{0} ({1}) ;\n".format(contact.get("contact").get("name"),
                                                       contact.get("contact").get("email"))\
                                for contact in contacts if contact.get("role") == "pointOfContact"]
            else:
                contacts_cct = ""

            # METADATA #
            md_created = dtparse(md.get("_created")).strftime("%a %d %B %Y (%Hh%M)")
            md_updated = dtparse(md.get("_modified")).strftime("%a %d %B %Y (%Hh%M)")

            # écriture des informations dans chaque colonne correspondante
            wbsheet.write(offset, 0, md.get("title"))
            wbsheet.write(offset, 1, md.get("name"))
            wbsheet.write(offset, 2, md.get("path"))
            wbsheet.write(offset, 3, " ; ".join(li_motscles))
            wbsheet.write(offset, 4, md.get("abstract"), style_wrap)
            wbsheet.write(offset, 5, " ; ".join(li_theminspire))
            wbsheet.write(offset, 6, md.get("type"))
            wbsheet.write(offset, 7, format_version)
            wbsheet.write(offset, 8, srs)
            wbsheet.write(offset, 9, md.get("features"))
            wbsheet.write(offset, 10, md.get("geometry"))
            wbsheet.write(offset, 11, owner)
            wbsheet.write(offset, 12, data_created.decode('latin1'))
            wbsheet.write(offset, 13, data_updated.decode('latin1'))
            wbsheet.write(offset, 14, md_created.decode('latin1'))
            wbsheet.write(offset, 15, md_updated.decode('latin1'))
            wbsheet.write(offset, 16, inspire_valid)
            wbsheet.write(offset, 17, len(contacts))
            wbsheet.write(offset, 18, contacts_cct, style_wrap)
            wbsheet.write(offset, 20, xlwt.Formula(link_visu), style_url)
            wbsheet.write(offset, 21, xlwt.Formula(link_edit), style_url)

        print(sorted(md.keys()))

        # end of function
        return

    # ------------ Setting workbook ---------------------

    def set_workbook_structure(self):
        """ TO DO
        """
        # styles
        style_header = xlwt.easyxf('pattern: pattern solid, fore_colour black;'
                                   'font: colour white, bold True, height 220;'
                                   'align: horiz center')
        style_url = xlwt.easyxf(u'font: underline single')
        style_wrap = xlwt.easyxf('align: wrap True')

        # sheets
        sheet_mds = book.add_sheet('Metadonnées', cell_overwrite_ok=True)

        # headers
        sheet_mds.write(0, 0, "Titre", style_header)
        sheet_mds.write(0, 1, "Nom de la ressource", style_header)
        sheet_mds.write(0, 2, "Emplacement", style_header)
        sheet_mds.write(0, 3, "Mots-clés", style_header)
        sheet_mds.write(0, 4, "Résumé", style_header)
        sheet_mds.write(0, 5, "Thématiques INPIRES", style_header)
        sheet_mds.write(0, 6, "Type", style_header)
        sheet_mds.write(0, 7, "Format", style_header)
        sheet_mds.write(0, 8, "SRS", style_header)
        sheet_mds.write(0, 9, "Nombre d'objets", style_header)
        sheet_mds.write(0, 10, "Géométrie", style_header)
        sheet_mds.write(0, 11, "Propriétaire", style_header)
        sheet_mds.write(0, 12, "Données - Création", style_header)
        sheet_mds.write(0, 13, "Données - Modification", style_header)
        sheet_mds.write(0, 14, "Métadonnées - Création", style_header)
        sheet_mds.write(0, 15, "Métadonnées - Modification", style_header)
        sheet_mds.write(0, 16, "Conformité INSPIRE", style_header)
        sheet_mds.write(0, 17, "# contacts", style_header)
        sheet_mds.write(0, 18, "Points de contacts", style_header)
        sheet_mds.write(0, 19, "Points de contacts", style_header)
        sheet_mds.write(0, 20, "Visualiser sur l'OpenCatalog", style_header)
        sheet_mds.write(0, 21, "Editer sur Isogeo", style_header)

        # columns width
        sheet_mds.col(0).width = 50 * 100
        sheet_mds.col(1).width = 40 * 256
        sheet_mds.col(4).width = 75 * 256

        # end of method
        return

    def set_worksheets(self, has_vector=1, has_raster=1, has_service=1, has_resource=1):
        """ adds news sheets depending on present metadata types
        """
        # SHEETS & HEADERS
        if has_vector:
            self.ws_vector = self.create_sheet(title="Vecteurs")
            # headers
            self.ws_vector.append([i for i in self.li_cols_vector])
        else:
            pass

        if has_raster:
            self.ws_raster = self.create_sheet(title="Raster")
            # headers
            self.ws_raster.append([i for i in self.li_cols_raster])
        else:
            pass

        if has_service:
            self.ws_services = self.create_sheet(title="Services")
            # headers
            self.ws_services.append([i for i in self.li_cols_service])
        else:
            pass

        if has_resource:
            self.ws_resources = self.create_sheet(title="Ressources")
            # headers
            self.ws_resources.append([i for i in self.li_cols_resource])
        else:
            pass

        # CLEAN UP & TUNNING
        for sheet in self.worksheets:
            # Freezing panes
            c = sheet['B2']
            sheet.freeze_panes = c

            # Print properties
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Others properties
            wsprops = sheet.sheet_properties
            wsprops.filterMode = True

        # end of method
        return

###############################################################################
###### Stand alone program ########
###################################

if __name__ == '__main__':
    """ Standalone execution and tests
    """
    # ------------ Specific imports ---------------------
    from ConfigParser import SafeConfigParser   # to manage options.ini
    from datetime import datetime
    from os import path

    # Custom modules
    from isogeo_sdk import Isogeo

    # ------------ Settings from ini file ----------------
    if not path.isfile(path.realpath(r"..\settings.ini")):
        print("ERROR: to execute this script as standalone, you need to store your Isogeo application settings in a isogeo_params.ini file. You can use the template to set your own.")
        import sys
        sys.exit()
    else:
        pass

    config = SafeConfigParser()
    config.read(r"..\settings.ini")

    settings = {s: dict(config.items(s)) for s in config.sections()}
    app_id = settings.get('auth').get('app_id')
    app_secret = settings.get('auth').get('app_secret')
    client_lang = settings.get('basics').get('def_codelang')

    # ------------ Connecting to Isogeo API ----------------
    # instanciating the class
    isogeo = Isogeo(client_id=app_id,
                    client_secret=app_secret,
                    lang="fr")

    token = isogeo.connect()

    # ------------ Isogeo search --------------------------
    search_results = isogeo.search(token,
                                   sub_resources=isogeo.sub_resources_available,
                                   preprocess=1)

    # ------------ REAL START ----------------------------
    wb = Isogeo2xlsx()
    wb.set_worksheets()

    # saving the test file
    dstamp = datetime.now()
    wb.save(r"..\output\test_isogeo2xlsx_{0}{1}{2}{3}{4}{5}.xlsx".format(dstamp.year,
                                                                         dstamp.month,
                                                                         dstamp.day,
                                                                         dstamp.hour,
                                                                         dstamp.minute,
                                                                         dstamp.second))




### DEV NOTES
# http://wiki.openstreetmap.org/wiki/FR:Parcourir#URL_avec_bbox
# http://wiki.openstreetmap.org/wiki/Layer_URL_parameter
# https://www.openstreetmap.org/?bbox=22.3418234%2C57.5129102%2C22.5739625%2C57.6287332&layers=H&box=yes
