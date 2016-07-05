# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ----------------------------------------------------------------------------
# Name:         Isogeo
# Purpose:      Get metadatas from an Isogeo share and store it into files
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      18/12/2015
# Updated:      22/01/2016
# ---------------------------------------------------------------------------

# ############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from ConfigParser import SafeConfigParser
from datetime import datetime
import logging      # log files
from logging.handlers import RotatingFileHandler
from os import listdir, path
from sys import argv, exit
from tkFileDialog import askopenfilename
from Tkinter import Tk, StringVar, IntVar, Image, PhotoImage   # GUI
from ttk import Label, Button, Entry, Checkbutton, Combobox  # advanced widgets
from ttk import Labelframe, Progressbar, Style  # advanced widgets

# 3rd party library
from openpyxl import load_workbook
from isogeo_pysdk import Isogeo

# Custom modules
from modules.isogeo2xlsx import Isogeo2xlsx
from modules.isogeo2docx import Isogeo2docx

# ############################################################################
# ########## Global ###############
# ##################################

# VERSION
_version = "1.1"

# LOG FILE ##
# see: http://sametmax.com/ecrire-des-logs-en-python/
logger = logging.getLogger()
logging.captureWarnings(True)
logger.setLevel(logging.DEBUG)  # all errors will be get
log_form = logging.Formatter('%(asctime)s || %(levelname)s || %(message)s')
logfile = RotatingFileHandler('isogeo2office.log', 'a', 5000000, 1)
logfile.setLevel(logging.DEBUG)
logfile.setFormatter(log_form)
logger.addHandler(logfile)
logger.info('\n============== Isogeo => Office =============')


# ############################################################################
# ########## Classes ###############
# ##################################

class Isogeo2office(Tk):
    """ UI Class to
    docstring for Isogeo to Office
    """
    # attributes and global actions
    logger.info('Version: {0}'.format(_version))

    def __init__(self, ui_launcher=1):
        """
        """
        #
        if not ui_launcher:
            self.no_ui_launcher()
        else:
            pass
        Tk.__init__(self)
        # ------------ Settings ----------------
        self.settings_load()
        app_id = self.settings.get('auth').get('app_id')
        app_secret = self.settings.get('auth').get('app_secret')
        client_lang = self.settings.get('basics').get('def_codelang')

        # ------------ Isogeo authentification ----------------
        self.isogeo = Isogeo(client_id=app_id,
                             client_secret=app_secret,
                             lang=client_lang,
                             platform="qa")
        self.token = self.isogeo.connect()

        # ------------ Isogeo search ----------------
        self.search_results = self.isogeo.search(self.token,
                                                 page_size=0,
                                                 whole_share=0)

        print(self.search_results.get("results"))

        # ------------ Variables ----------------
        li_tpls = [path.abspath(path.join(r'templates', tpl))
                   for tpl in listdir(r'templates')
                   if path.splitext(tpl)[1].lower() == ".docx"]

        # ------------ UI ----------------
        self.title('isogeo2office - ToolBox')
        icon = Image("photo", file=r'img/favicon_isogeo.gif')
        self.call('wm', 'iconphoto', self._w, icon)
        self.style = Style().theme_use('clam')

        # Frames
        fr_global = Labelframe(self,
                               name='global',
                               text="Générique")

        fr_excel = Labelframe(self,
                              name='excel',
                              text="Export Excel")

        fr_word = Labelframe(self,
                             name='word',
                             text="Export Word")

        # ## GLOBAL ##
        url_input = StringVar(self)
        url_input.set(self.settings.get('basics').get('def_oc'))
        # logo
        logo_isogeo = PhotoImage(file=r'img/logo_isogeo.gif')
        Label(self,
              borderwidth=2,
              image=logo_isogeo).grid(row=1, rowspan=2,
                                      column=0, padx=2,
                                      pady=2, sticky="W")

        lb_count_avail_resources = Label(fr_global,
                                         text="{} métadonnées partagées"\
                                              .format(self.search_results.get('total'))).pack()

        # Frame: Progression bar
        self.FrProg = Labelframe(self,
                                 name='progression',
                                 text=self.blabla.get('gui_prog'))
        # variables
        self.status = StringVar(self.FrProg, '')
        # widgets
        self.prog_layers = Progressbar(self.FrProg,
                                       orient="horizontal")
        Label(master=self.FrProg,
              textvariable=self.status,
              foreground='DodgerBlue').pack()
        # widgets placement
        self.prog_layers.pack(expand=1, fill='both')

        # OpenCatalog URL
        lb_input_oc = Label(fr_global,
                            text="Coller l'URL d'un OpenCatalog").pack()
        self.ent_opencatalog = Entry(fr_global,
                                     textvariable=url_input,
                                     width=100)
        self.ent_opencatalog.pack()
        self.ent_opencatalog.insert(0, "yhouh")
        # self.settings.get('basics').get('def_oc')
        self.ent_opencatalog.focus_set()

        fr_global.grid(row=1, sticky="WE")

        # ------------------------------------------------------------

        # ## EXCEL ##
        # variables
        output_xl = StringVar(self)
        self.opt_xl_join = IntVar(fr_excel)
        self.input_xl = ""
        li_input_xl_cols = []
        self.input_xl_join_col = StringVar()

        # output file
        lb_output_xl = Label(fr_excel,
                             text="Nom du fichier en sortie: ").pack()
        ent_output_xl = Entry(fr_excel,
                              text="Nom du fichier en sortie: ",
                              textvariable=output_xl,
                              width=50).pack()

        # matching with another Excel file
        self.fr_input_xl_join = Labelframe(fr_excel,
                                           name='excel_joiner',
                                           text="Jointure à partir d'un autre tableur Excel")
        caz_xl_join = Checkbutton(fr_excel,
                                  text=u'Joindre avec un autre fichier Excel',
                                  variable=self.opt_xl_join,
                                  command=lambda: self.ui_switch_xljoiner())
        caz_xl_join.pack()

        bt_browse_input_xl = Button(self.fr_input_xl_join,
                                    text="Choisir un fichier en entrée",
                                    command=lambda: self.get_input_xl()).pack()
        lb_input_xl = Label(self.fr_input_xl_join,
                            text=self.input_xl).pack()

        cb_input_xl_cols = Combobox(self.fr_input_xl_join,
                                    textvariable=self.input_xl_join_col,
                                    values=li_input_xl_cols,
                                    width=100)
        cb_input_xl_cols.pack()


        # TO COMPLETE
        # self.fr_input_xl_join.pack()

        Button(fr_excel,
               text="Excelization !",
               command=lambda: self.process_excelization(output_xl)).pack()

        fr_excel.grid(row=2)

        # ------------------------------------------------------------

        # ## WORD ##
        # variables
        self.tpl_input = StringVar(self)
        # pick a template
        lb_input_tpl = Label(fr_word,
                             text="Choisir un template").pack()
        cb_available_tpl = Combobox(fr_word,
                                    textvariable=self.tpl_input,
                                    values=li_tpls,
                                    width=100)
        cb_available_tpl.pack()

        Button(fr_word,
               text="Wordification !",
               command=lambda: process_wordification()).pack()
        # packing frame
        fr_word.grid(row=3)

# ----------------------------------------------------------------------------

    def get_input_xl(self):
        """ Get the path of the input Excel file with a browse dialog
        """
        self.input_xl = askopenfilename(parent=self,
                                        filetypes=[("Excel 2010 files","*.xlsx"),("Excel 2003 files","*.xls")],
                                        title=u"Choisir le fichier Excel à partir duquel faire la jointure")

        # testing file choosen
        if self.input_xl:
            print(self.input_xl)
            pass
        elif path.splittext(self.input_xl)[1] != ".xlsx":
            print("Pas le bon format")
        else:
            print(u'Aucun fichier sélectionné')
            return

        # get headers names
        xlsx_in = load_workbook(filename=self.input_xl,
                                read_only=True,
                                guess_types=True,
                                use_iterators=True)
        ws1 = xlsx_in.worksheets[0]  # ws = première feuille
        cols_names = [ws1.cell(row=ws1.min_row, column=col).value for col in range(1, ws1.max_column)]

        # end of method
        return

    def ui_switch_xljoiner(self):
        """ Enable/disable the form for input xl to join.
        """
        if self.opt_xl_join.get():
            self.fr_input_xl_join.pack()
        else:
            self.fr_input_xl_join.pack_forget()
        # end of function
        return


    def get_basic_metrics(self):
        """ TO DO
        """
        empty_search = self.isogeo.search(self.token,
                                          # query="keyword:isogeo:2015",
                                          page_size=0,
                                          whole_share=0,
                                          prot='http')

        # end of method
        return len(empty_search.get('results'))


    def get_search_results(self):
        """ TO DO
        """
        pass

# ----------------------------------------------------------------------------

    def settings_load(self):
        """ TO DO
        """
        config = SafeConfigParser()
        config.read(r"settings.ini")
        self.settings = {s:dict(config.items(s)) for s in config.sections()}

        # end of method
        return

    def settings_save(self, ):
        """ TO DO
        """

        # end of method
        return

# ----------------------------------------------------------------------------

    def get_url_base(self, url_input):
        """ TO DO
        """
        # get the OpenCatalog URL given
        if not url_input[-1] == '/':
            url_input = url_input + '/'
        else:
            pass

        # get the clean url
        url_output = url_input[0:url_input.index(url_input.rsplit('/')[6])]

        # end of method
        return url_output

# ----------------------------------------------------------------------------

    def process_excelization(self, output_filename):
        """ TO DO
        """
        includes = ["conditions",
                    "contacts",
                    "coordinate-system",
                    "events",
                    "feature-attributes",
                    "keywords",
                    "limitations",
                    "links",
                    "specifications"]

        self.search_results = self.isogeo.search(self.token,
                                                 page_size=0,
                                                 whole_share=0)

        # ------------ REAL START ----------------------------
        wb = Isogeo2xlsx()
        wb.set_worksheets()

        # parsing metadata
        for md in search_results.get('results'):
            wb.store_metadatas(md)

        # tunning
        wb.tunning_worksheets()

        # saving the test file
        dstamp = datetime.now()
        wb.save(r"output\{0}.xlsx".format())

        # end of method
        return

    def process_wordification(self, search_results):
        """ TO DO
        """
        for md in search_results.get("results"):
            tpl = DocxTemplate(path.realpath(self.tpl_input.get()))
            toDocx.md2docx(tpl, md, url_oc)
            dstamp = datetime.now()
            if not md.get('name'):
                md_name = "NR"
            elif '.' in md.get('name'):
                md_name = md.get("name").split(".")[1]
            else:
                md_name = md.get("name")
            tpl.save(r"..\output\{0}_{8}_{7}_{1}{2}{3}{4}{5}{6}.docx".format("TestDemoDev",
                                                                             dstamp.year,
                                                                             dstamp.month,
                                                                             dstamp.day,
                                                                             dstamp.hour,
                                                                             dstamp.minute,
                                                                             dstamp.second,
                                                                             md.get("_id")[:5],
                                                                             md_name))
            del tpl

        # end of method
        return

# ----------------------------------------------------------------------------

    def no_ui_launcher(self):
        """ Execute the scripts without displaying the UI and using
        settings.ini
        """
        logger.info('Launched from command prompt')
        exit()
        pass

# ###############################################################################
# ###### Stand alone program ########
# ###################################

if __name__ == '__main__':
    """ standalone execution
    """
    if len(argv) < 2:
        app = Isogeo2office(ui_launcher=1)
        app.mainloop()
    elif argv[1] == str(1):
        print("launch UI")
        app = Isogeo2office(ui_launcher=1)
        app.mainloop()
    else:
        print("launch without UI")
        app = Isogeo2office(ui_launcher=0)
